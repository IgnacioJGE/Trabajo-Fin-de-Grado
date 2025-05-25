

import json
import os
import string
from dotenv import load_dotenv
import requests
import openpyxl


def decidir_persona_asignada(tipo_solicitud):#toca cambiar esto para que el else sea Joseluis
    if tipo_solicitud == "Padua":
        return "Paco"
    elif tipo_solicitud in ["Equipos, etiquetadoras e impresoras", "Viaweb", "Otros"]:
        return "Jose Luis Martín"
    elif tipo_solicitud == ["Albarán digital", "Clientes"]:
        return "Jeickon"
    elif tipo_solicitud == "Desarrollos y proyectos":
        return "Álvaro"
    else:
        return "Jose Luis Martín"
    

def calcular_importancia(importancia):
   dicc_valores={
       "Crítica": 5,
       "Importante": 4,
       "Algo importante" : 3,
       "Poco importante": 2,
       "Mínima": 1
   }
   dicc_inverso={
         5: "Crítica",
         4: "Importante",
         3: "Algo importante",
         2: "Poco importante",
         1: "Mínima"
   }
   valor= dicc_valores.get(importancia)
   valor:int=(valor+3)/2
   return dicc_inverso.get(int(valor))


async def  token_auth():
    load_dotenv()
    client_id = os.getenv("client_id")
    client_secret = os.getenv("client_secret")
    tenant_id = os.getenv("tenant_id")
    tenant=os.getenv("tenant")
    url1 = os.getenv("url1")
    
    data = {
        'grant_type':'client_credentials',
        'resource': "00000003-0000-0ff1-ce00-000000000000/" + tenant + ".sharepoint.com@" + tenant_id, 
        'client_id': client_id,
        'client_secret': client_secret,
    }   
    headers = {
        'Content-Type':'application/x-www-form-urlencoded'
    }
    r = requests.post(url1, data=data, headers=headers)
    json_data = json.loads(r.text)
    return json_data


def buscarcontra(correo:string,idpadua:string):
    load_dotenv()
    correoencontrado:bool = False
    idEncontrado:bool = False
    excel1ruta=os.getenv("contra1")
    excel2ruta=os.getenv("contra2")
    excelsheet1 = openpyxl.load_workbook(excel1ruta)
    for sheet in excelsheet1.sheetnames:
        worksheet = excelsheet1[sheet]
        for row in worksheet.iter_rows():
            if row[0].value == idpadua and row[5].value != correo:
                idEncontrado = True
                return("id")
            if row[0].value != idpadua and row[5].value == correo:
                print(row[0].value)
                print(row[5].value)
                correoencontrado = True
                return("correo")
                
            if row[0].value == idpadua and row[5].value == correo:
                return(row[7].value)
            
    excelsheet2 = openpyxl.load_workbook(excel2ruta)       
    if correoencontrado!=True and idEncontrado!=True:
        for sheet in excelsheet2.sheetnames:
            worksheet = excelsheet2[sheet]
            for row in worksheet.iter_rows():
                if row[0].value == idpadua and row[5].value != correo:
                    idEncontrado = True
                    return("id")
                if row[0].value != idpadua and row[5].value == correo:
                    correoencontrado = True
                    return("correo")
                if row[0].value == idpadua and row[5].value == correo:
                    return(row[7].value) 
    